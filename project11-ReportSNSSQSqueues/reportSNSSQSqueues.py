import boto3
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

# Create session with your desired profile
session = boto3.Session() #to use with cloudshell
#if running the code via the local machine, first login with the command saml2aws login
# then deactivate L7 and activate that line: session = boto3.Session(profile_name="saml", region_name="eu-central-1") 

sqs_client = session.client("sqs")
sns_client = session.client("sns")

sqs_urls = sqs_client.list_queues().get('QueueUrls', [])
report_rows = []

for sqs_url in sqs_urls:
    queue_attributes = sqs_client.get_queue_attributes(
        QueueUrl=sqs_url,
        AttributeNames=["QueueArn", "RedrivePolicy"]
    )
    queue_arn = queue_attributes["Attributes"]["QueueArn"]
    redrive_policy = queue_attributes["Attributes"].get("RedrivePolicy", "n.A.")

    # Flag to check if this queue has any SNS subscription
    has_subscription = False

    next_token = None
    while True:
        kwargs = {"NextToken": next_token} if next_token else {}
        response = sns_client.list_subscriptions(**kwargs)

        for sub in response["Subscriptions"]:
            if sub["Protocol"] == "sqs" and sub["Endpoint"] == queue_arn:
                has_subscription = True

                # Get subscription filter policy
                try:
                    sub_attrs = sns_client.get_subscription_attributes(
                        SubscriptionArn=sub["SubscriptionArn"]
                    )
                    filter_policy = sub_attrs["Attributes"].get("FilterPolicy", "n.A.")
                except Exception:
                    filter_policy = "n.A."

                report_rows.append({
                    "SQS_Queue_URL": sqs_url,
                    "SQS_Queue_ARN": queue_arn,
                    "SNS_Topic_ARN": sub["TopicArn"],
                    "Subscription_ARN": sub["SubscriptionArn"],
                    "Subscription_Filter_Policy": filter_policy,
                    "Redrive_Policy": redrive_policy
                })

        next_token = response.get("NextToken")
        if not next_token:
            break

    # If no subscription found, add a row with empty subscription columns
    if not has_subscription:
        report_rows.append({
            "SQS_Queue_URL": sqs_url,
            "SQS_Queue_ARN": queue_arn,
            "SNS_Topic_ARN": "n.A",
            "Subscription_ARN": "n.A",
            "Subscription_Filter_Policy": "n.A.",
            "Redrive_Policy": redrive_policy
        })

# Save to Excel
excel_file = "test2sqs_sns_report.xlsx"
df = pd.DataFrame(report_rows)
df.to_excel(excel_file, index=False)

# Adjust column widths
wb = load_workbook(excel_file)
ws = wb.active

for col_idx, col in enumerate(df.columns, 1):
    max_length = max(
        df[col].astype(str).map(len).max(),
        len(col)
    )
    adjusted_width = max_length + 2
    ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

wb.save(excel_file)

print(f"Done! Report saved to: {excel_file}")
